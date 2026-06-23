#!/usr/bin/env python3
"""
BC端库存报表自动生成脚本。

用法：
    python build_report.py

流程：
    1. 扫描 data/input/ 下的 Excel 文件
    2. 按子表名识别数据类型（不依赖文件名）
    3. 从模板读取目标列清单，按表头名称匹配重排
    4. 打印校验报告，等待用户确认
    5. 写入模板副本，自动填充公式
    6. 输出到 data/output/，原始文件归档
"""

import openpyxl
import openpyxl.worksheet.formula as fml
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from pathlib import Path
import shutil
from datetime import datetime, timedelta
from collections import defaultdict

# ============================================================
# 常量
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "data" / "input"
TEMPLATE_PATH = BASE_DIR / "data" / "template" / "BC端库存报表模板.xlsx"
OUTPUT_DIR = BASE_DIR / "data" / "output"
ARCHIVE_DIR = BASE_DIR / "data" / "archive"

# 导出子表名 → 内部标识
SHEET_TYPE_MAP = {
    "收货记录表": "inbound",
    "库存汇总查询": "inventory",
    "出库全流程记录": "outbound",
}

# 每类数据的配置：模板中的(子表名, 目标列清单来源子表, 写入起始列, 公式列)
TYPE_CONFIG = {
    "inbound": {
        "target_sheet": "收货记录查询",
        "column_source_sheet": "收货记录查询高级筛选",
        "column_source_range": "A",       # 从 A 列开始读目标列名
        "write_start_col": 5,             # E 列 = 5
        "formula_cols": [1, 2, 3, 4],     # A-D 列
    },
    "inventory": {
        "target_sheet": "总库存",
        "column_source_sheet": "总库存",
        "column_source_range": "F",        # 从 F 列开始读目标列名
        "write_start_col": 6,              # F 列 = 6
        "formula_cols": [1, 2, 3, 4, 5],  # A-E 列
    },
    "outbound": {
        "target_sheet": "出库全流程",
        "column_source_sheet": "出库全流程高级筛选",
        "column_source_range": "A",
        "write_start_col": 5,              # E 列 = 5
        "formula_cols": [1, 2, 3, 4],     # A-D 列
    },
}


# ============================================================
# 扫描 & 识别
# ============================================================

def scan_input_files(input_dir: Path) -> list[Path]:
    """扫描 input 目录，返回所有 xlsx/xls 文件路径列表。"""
    files = []
    for pattern in ["*.xlsx", "*.xls"]:
        files.extend(input_dir.glob(pattern))
    # 排除临时文件（~$ 开头）
    files = [f for f in files if not f.name.startswith("~$")]
    return sorted(files)


def identify_data_type(wb: openpyxl.Workbook) -> str | None:
    """根据工作簿的子表名识别数据类型。

    遍历 SHEET_TYPE_MAP，只要工作簿中存在匹配的子表即返回对应类型。
    返回 None 表示无法识别。
    """
    sheet_names = set(wb.sheetnames)
    for sheet_name, dtype in SHEET_TYPE_MAP.items():
        if sheet_name in sheet_names:
            return dtype
    return None


# ============================================================
# 列映射 & 重排
# ============================================================

def read_target_columns(wb: openpyxl.Workbook, dtype: str) -> list[str]:
    """从模板读取目标列清单。

    读取规则：
    - inbound: 收货记录查询高级筛选 的 A-AG 列（读到第一个空列为止）
    - inventory: 总库存 的 F-W 列
    - outbound: 出库全流程高级筛选 的 A-BO 列（读到第一个空列为止）

    返回列名列表，顺序与模板一致。
    """
    config = TYPE_CONFIG[dtype]
    source_sheet = config["column_source_sheet"]
    source_range = config["column_source_range"]
    ws = wb[source_sheet]

    # 确定起始列号
    start_col = openpyxl.utils.column_index_from_string(source_range)

    columns = []
    for c in range(start_col, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val is None:
            # 对于 A 列起始的，遇到空列停止；对于 F 列起始的，读到 W 列位置后如果为空也停止
            break
        columns.append(str(val).strip())

    return columns


def match_and_reorder(
    source_ws: openpyxl.worksheet.worksheet.Worksheet,
    target_columns: list[str],
) -> tuple[list[list], dict]:
    """将源子表列按目标列清单匹配并重排。

    参数:
        source_ws: 导出文件的数据子表
        target_columns: 目标列名清单（有序）

    返回:
        (reordered_data, match_report)
        reordered_data: list[list] 重排后的数据行（每行长度 = len(target_columns)）
        match_report: {
            "matched": int,       # 匹配成功列数
            "missing": list[str], # 目标有但导出无的列名
            "extra": list[str],   # 导出有但目标无的列名
        }
    """
    # 读取导出文件的表头（第1行）
    source_headers = []
    header_col_map = {}  # 列名 → 列号(1-based)
    for c in range(1, source_ws.max_column + 1):
        val = source_ws.cell(row=1, column=c).value
        if val is not None:
            name = str(val).strip()
            source_headers.append(name)
            header_col_map[name] = c

    # 建立映射：目标列名 → 源文件列号(None 表示缺失)
    col_mapping: list[int | None] = []
    missing = []
    for tc in target_columns:
        if tc in header_col_map:
            col_mapping.append(header_col_map[tc])
        else:
            col_mapping.append(None)
            missing.append(tc)

    # 导出中有但目标清单没有的列
    target_set = set(target_columns)
    extra = [h for h in source_headers if h not in target_set]

    # 重排数据行（跳过表头行 1）
    reordered_data = []
    for r in range(2, source_ws.max_row + 1):
        row = []
        has_data = False
        for src_col in col_mapping:
            if src_col is not None:
                val = source_ws.cell(row=r, column=src_col).value
                row.append(val)
                if val is not None:
                    has_data = True
            else:
                row.append(None)
        if has_data:  # 跳过全空行
            reordered_data.append(row)

    return reordered_data, {
        "matched": len(target_columns) - len(missing),
        "missing": missing,
        "extra": extra,
    }


# ============================================================
# 校验报告模块
# ============================================================

def generate_validation_report(results: dict) -> str:
    """根据收集的数据生成校验报告文本。

    参数:
        results: {
            dtype: {
                "files": [Path, ...],
                "total_rows": int,
                "match_report": dict,      # 来自 match_and_reorder
                "data": list[list],        # 重排后的数据
            }
        }

    返回: 格式化的中文校验报告字符串
    """
    lines = []
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    lines.append(f"{'='*50}")
    lines.append(f"  数据校验报告 ({yesterday})")
    lines.append(f"{'='*50}")

    type_names = {
        "inbound": "收货记录表",
        "inventory": "库存汇总查询",
        "outbound": "出库全流程记录",
    }

    for dtype in ["inbound", "inventory", "outbound"]:
        if dtype not in results:
            lines.append(f"\n【{type_names[dtype]}】❌ 未找到数据")
            continue

        r = results[dtype]
        files = r["files"]
        data = r["data"]
        mr = r["match_report"]
        name = type_names[dtype]

        lines.append(f"\n【{name}】({len(files)} 个文件，共 {len(data)} 行)")

        # 文件名列表
        for f in files:
            lines.append(f"  📄 {f.name}")

        # 列匹配结果
        total_target = mr["matched"] + len(mr["missing"])
        if mr["missing"]:
            lines.append(f"  列匹配: ✅ {mr['matched']}/{total_target}  ⚠️ 缺失: {', '.join(mr['missing'])}")
        else:
            lines.append(f"  列匹配: ✅ {total_target}/{total_target} 全部匹配")
        if mr["extra"]:
            lines.append(f"  ℹ️ 导出多出的列(已丢弃): {', '.join(mr['extra'])}")

        # 日期范围（尝试取第1列中的日期信息）
        if data:
            first_vals = [row[0] for row in data if row[0] is not None]
            if first_vals:
                lines.append(f"  首行时间: {str(first_vals[0])[:19]}")
                lines.append(f"  末行时间: {str(first_vals[-1])[:19]}")

        # 空值统计
        if data:
            null_cols = []
            for ci in range(len(data[0])):
                null_count = sum(1 for row in data if row[ci] is None or str(row[ci]).strip() == "")
                if null_count == len(data):
                    # 全列为空
                    col_name = mr.get("target_names", [f"列{ci}"])[ci] if ci < len(mr.get("target_names", [])) else f"列{ci}"
                    null_cols.append(f"{col_name}(全空)")
            if null_cols:
                lines.append(f"  空值: {', '.join(null_cols[:5])}")

    lines.append(f"\n{'='*50}")
    return "\n".join(lines)


def extract_date_range(data: list[list], col_idx: int = 0) -> tuple[str, str] | None:
    """尝试从数据中提取日期范围。"""
    first = None
    last = None
    for row in data:
        val = row[col_idx] if col_idx < len(row) else None
        if val is not None and str(val).strip():
            last = str(val).strip()[:19]
            if first is None:
                first = last
    if first:
        return (first, last)
    return None


# ============================================================
# 数据写入与公式填充
# ============================================================

def write_data_to_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    data: list[list],
    start_col: int,
) -> int:
    """将重排后的数据写入子表。

    从第 2 行开始写入，start_col 为数据起始列号(1-based)。
    返回写入的行数。
    """
    for ri, row_data in enumerate(data):
        target_row = ri + 2  # 第1行是表头，从第2行开始写
        for ci, val in enumerate(row_data):
            ws.cell(row=target_row, column=start_col + ci, value=val)

    return len(data)


def auto_fill_formulas(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    formula_cols: list[int],
    data_row_count: int,
) -> int:
    """将第2行公式向下填充至所有数据行。

    参数:
        ws: 目标子表
        formula_cols: 需要填充的列号列表(1-based)，如 [1,2,3,4]
        data_row_count: 数据总行数

    返回: 实际填充的公式列数
    """
    filled = 0
    last_row = 1 + data_row_count  # 表头1行 + N行数据

    for c in formula_cols:
        src_formula = ws.cell(row=2, column=c).value
        if src_formula is None:
            continue

        # 处理 ArrayFormula 对象（动态数组公式如 XLOOKUP）
        if isinstance(src_formula, fml.ArrayFormula):
            src_str = src_formula.text  # 去掉花括号，取纯公式文本
        else:
            src_str = str(src_formula)

        if not src_str.startswith("="):
            continue

        col_letter = get_column_letter(c)
        origin = f"{col_letter}2"

        for r in range(2, last_row + 1):
            target_cell = f"{col_letter}{r}"
            try:
                translated = Translator(src_str, origin=origin).translate_formula(target_cell)
                ws.cell(row=r, column=c).value = translated
            except Exception:
                # 翻译失败时直接赋原公式字符串（Excel 打开时会自动调整）
                ws.cell(row=r, column=c).value = src_str
        filled += 1

    return filled


# ============================================================
# 归档模块
# ============================================================

def archive_input_files(files: list[Path], archive_subdir: Path) -> None:
    """将处理完的原始文件移动到归档目录。"""
    archive_subdir.mkdir(parents=True, exist_ok=True)
    for f in files:
        dst = archive_subdir / f.name
        # 如果同名文件已存在，加序号
        if dst.exists():
            stem = f.stem
            suffix = f.suffix
            for i in range(1, 100):
                dst = archive_subdir / f"{stem}_{i}{suffix}"
                if not dst.exists():
                    break
        try:
            shutil.move(str(f), str(dst))
            print(f"  已归档: {f.name}")
        except Exception:
            print(f"  ⚠️ 归档失败: {f.name}")


# ============================================================
# 主入口
# ============================================================

def main():
    """主流程。"""
    print("BC端库存报表自动生成工具")
    print("-" * 40)

    # 1. 检查模板
    if not TEMPLATE_PATH.exists():
        print(f"❌ 模板文件不存在: {TEMPLATE_PATH}")
        print("   请先运行: python scripts/prepare_template.py")
        return

    # 2. 扫描输入文件
    input_files = scan_input_files(INPUT_DIR)
    if not input_files:
        print(f"❌ data/input/ 中没有 Excel 文件，请先放入导出的原始文件。")
        return

    print(f"\n📂 扫描到 {len(input_files)} 个文件:")
    for f in input_files:
        print(f"  - {f.name}")

    # 3. 识别每个文件的数据类型，读取并重排数据
    results = {}  # dtype → {files, data, match_report}
    all_files_to_archive = []

    # 预加载模板（一次性，不在文件循环内重复加载）
    template_wb = openpyxl.load_workbook(TEMPLATE_PATH)

    for fpath in input_files:
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
        except Exception as e:
            print(f"⚠️ 无法打开 {fpath.name}: {e}")
            continue

        dtype = identify_data_type(wb)
        if dtype is None:
            print(f"⚠️ 未识别: {fpath.name}（子表名不匹配）")
            wb.close()
            continue

        print(f"  ✅ {fpath.name} → {dtype}")

        # 读取导出数据子表
        source_sheet_name = None
        for sn, dt in SHEET_TYPE_MAP.items():
            if dt == dtype and sn in wb.sheetnames:
                source_sheet_name = sn
                break

        if source_sheet_name is None:
            print(f"    ⚠️ 找不到预期子表，跳过")
            wb.close()
            continue

        source_ws = wb[source_sheet_name]

        # 读取目标列清单（使用预加载的模板）
        target_columns = read_target_columns(template_wb, dtype)

        # 匹配并重排
        reordered_data, match_report = match_and_reorder(source_ws, target_columns)
        match_report["target_names"] = target_columns  # 保留列名供报告使用

        # 累积到 results
        if dtype not in results:
            results[dtype] = {
                "files": [],
                "data": [],
                "match_report": match_report,
            }
        results[dtype]["files"].append(fpath)
        results[dtype]["data"].extend(reordered_data)
        all_files_to_archive.append(fpath)

        wb.close()

    template_wb.close()

    # 更新行数
    for dtype in results:
        results[dtype]["total_rows"] = len(results[dtype]["data"])

    # 4. 打印校验报告
    report = generate_validation_report(results)
    print(report)

    # 5. 等待确认
    confirm = input("\n→ 确认生成报表? (y/n): ").strip().lower()
    if confirm != "y":
        print("已取消。")
        return

    # 6. 写入模板副本（直接打开模板，数据写入后另存到 output，不修改模板本身）
    print("\n📝 正在生成报表...")
    output_wb = openpyxl.load_workbook(TEMPLATE_PATH)

    for dtype, config in TYPE_CONFIG.items():
        if dtype not in results or not results[dtype]["data"]:
            continue

        target_sheet = config["target_sheet"]
        ws = output_wb[target_sheet]

        # 清空第2行起的所有旧数据（模板只有1行公式，但我们可能之前已写入）
        if ws.max_row > 1:
            for r in range(2, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).value = None

        data = results[dtype]["data"]
        start_col = config["write_start_col"]
        formula_cols = config["formula_cols"]

        # 写入数据
        written = write_data_to_sheet(ws, data, start_col)
        # 填充公式
        filled = auto_fill_formulas(ws, formula_cols, written)

        print(f"  [{target_sheet}] 写入 {written} 行，填充 {filled} 列公式")

    # 7. 保存输出
    yesterday = datetime.now() - timedelta(days=1)
    out_name = f"BC端库存报表_{yesterday.strftime('%m%d')}.xlsx"
    out_path = OUTPUT_DIR / out_name
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    try:
        output_wb.save(out_path)
    except Exception:
        print(f"\n❌ 写入失败，请检查输出文件是否被 Excel 占用")
        output_wb.close()
        return
    output_wb.close()
    print(f"\n✅ 报表已生成: {out_path}")

    # 8. 归档原始文件
    archive_subdir = ARCHIVE_DIR / yesterday.strftime("%Y-%m-%d")
    print(f"\n📦 归档原始文件到: {archive_subdir}")
    archive_input_files(all_files_to_archive, archive_subdir)

    print("\n🎉 完成!")


if __name__ == "__main__":
    main()
